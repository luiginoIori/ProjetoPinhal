import openpyxl
import matplotlib.pyplot as plt
import webbrowser
import locale
import matplotlib.image as mpimg
import os
import csv


locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
# Caminho do arquivo Excel
arquivo_excel = 'Energy - Orçado x Realizado.xlsx'

# Abre o arquivo
wb = openpyxl.load_workbook(arquivo_excel,data_only=True)

# Seleciona a aba (sheet) desejada
aba = wb.active  # ou wb['NomeDaAba']

# Seleciona pela linha e coluna (linha 1, coluna 1)

ultima_linha = aba.max_row
ultima_coluna = aba.max_column

meses = []
posicoes =  []
for i in range(ultima_coluna):
    celula = aba.cell(row=i+1, column=1).value
    posicoes.append(celula)
    posicoes.append(i+1)
    
def dados(pos):
    Realizadas =[]    
    orcado =[0]
    for i in range(ultima_linha):
        celula = aba.cell(row=1, column=i+1)    
        if celula.value == "Realizado":
            celula = aba.cell(row=pos, column=i+1).value 
            if celula is None:
                celula = 0  
            if celula < 0:         
                Realizadas.append(int(celula)*-1)  
            else:
                Realizadas.append(int(celula))
            
    for i in range(ultima_linha):
        celula = aba.cell(row=1, column=i+1)    
        if celula.value == "Orçado":
            celula = aba.cell(row=pos, column=i+1).value
            if celula is None:
                celula = 0       
            if celula < 0:
                orcado.append(int(celula)*-1)
            else:
                orcado.append(int(celula))
                
    return Realizadas,orcado      

descricoes = ['Entradas','Receitas','Aporte','Despesas','OPEX','CAPEX','Impostos']
resumo = [] 
for j in descricoes:
    linha = j
    
    for i in range(len(posicoes)): 
        if linha == posicoes[i]:
            pos = posicoes[i+1]   
            Entradas_Real = dados(pos)[0]
            Entradas_Orca = dados(pos)[1]             
            labels = []
            razao = 0
            # Soma dos valores
             
            if linha == 'Despesas' or linha == 'OPEX' or linha == 'CAPEX'  or linha == 'Aporte' or linha == 'Impostos' or linha == 'Receitas' or linha == 'Entradas':
                soma_real = sum(Entradas_Real)
                soma_orca = sum(Entradas_Orca)

                # Evita divisão por zero
                if soma_real != 0:
                    razao =  soma_real /soma_orca
                    resumo.append([linha,soma_orca,soma_real,razao])
                else:
                    razao = 0

                print(f"Soma Entradas Realizadas: {soma_real}")
                print(f"Soma Entradas Orçadas: {soma_orca}")
                print(f"Entradas Orçadas / Realizadas: {razao:.2f} " + linha)
            

            for i in range(len(Entradas_Real)):
                if i <= 2:
                    labels.append(f'{i+10}-24')
                if i > 2:
                    labels.append(f'{i-2}-25')
            
            x = range(len(labels))
            print(x)
            largura = 0.35
            
          # ...existing code...
            plt.figure(figsize=(10,6))
            barras_real = plt.bar([i - largura/2 for i in x], Entradas_Real, width=largura, label= linha +' Realizadas', alpha=0.5)            
            plt.bar([i + largura/2 for i in x], Entradas_Orca, width=largura, label= linha +' Orçadas', alpha=0.5)

            # Adiciona o valor formatado em cada barra de Entradas_Real
            for idx, barra in enumerate(barras_real):
                valor = Entradas_Real[idx]
                valor_formatado = locale.format_string('%.0f', valor, grouping=False)
                plt.text(barra.get_x() + barra.get_width()/2, barra.get_height(), valor_formatado, 
                        ha='center', va='bottom', fontsize=8, color='navy')

          
            #if razao != 0:
            try:
                plt.text(
                    0.5, 0.97,
                            'Total realizado: ' + locale.format_string('%.0f', soma_real, grouping=True),
                            ha='center', va='center', transform=plt.gca().transAxes, fontsize=12, color='navy', fontweight='bold'
                        )
            except NameError:
                pass
            try:
                plt.text(
                    0.5, 0.93,
                    'Total orçado: ' + locale.format_string('%.0f', soma_orca, grouping=True),
                    ha='center', va='center', transform=plt.gca().transAxes, fontsize=12, color='green', fontweight='bold'
                )
            except NameError:
                pass
            plt.text(
                0.5, 0.89,
                'Porcentagem executada no período em relação ao previsto: ' + str(int(razao * 100)) + '%',
                ha='center', va='center', transform=plt.gca().transAxes, fontsize=10, color='gray'
            )
                
            plt.xticks(x, labels)
            plt.xlabel('Meses')
            plt.ylabel('Valores')
            plt.title(linha + ': Realizadas x Orçadas')
            plt.legend()
            try:
                logo = mpimg.imread('D:\Controle Energy\logo_be.png')
                # Adiciona o logotipo no canto superior direito
                plt.figimage(logo, xo=plt.gcf().bbox.xmax - logo.shape[1] - 700, yo=plt.gcf().bbox.ymax - logo.shape[0] -130, zorder=100, alpha=0.20)
            except FileNotFoundError:
                pass  # Não adiciona logotipo se não encontrar o arquivo

            plt.tight_layout()
            plt.savefig(f'Imagens/{linha}.png')
            #plt.show()
            plt.close()
            
            # Gráficos de linha
            try:
                plt.figure(figsize=(8,4))
                plt.plot(labels, Entradas_Real, marker='o', color='navy', label='Realizadas')
                plt.plot(labels, Entradas_Orca, marker='o', color='green', label='Orçadas')
                plt.title(f'Gráfico de Linha - {linha}')
                plt.xlabel('Meses')
                plt.ylabel('Valores')
                plt.legend()
                plt.grid(True)
                plt.tight_layout()
                plt.savefig(f'Imagens/{linha}_linha.png')
                plt.close()
            except Exception as e:
                print(f'Erro ao gerar gráfico de linha para {linha}:', e)

#nome = 'graficos.html'
#webbrowser.open(f'd:\Controle Energy\{nome}')
with open('resumo.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(['Categoria', 'Total Orçado', 'Total Realizado', 'Razao'])
    writer.writerows(resumo)
'''
nome = 'index.html'
html_path = os.path.abspath(nome)
print(html_path)
webbrowser.open(f'file:///{html_path}')
            
   '''
