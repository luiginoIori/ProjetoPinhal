import subprocess
import streamlit as st
from PIL import Image
import os
import csv
import glob
import datetime
import base64
import openpyxl
import matplotlib.pyplot as plt
import webbrowser
import locale
import matplotlib.image as mpimg
import os
import csv
import plotly.graph_objects as go

descricoes = ['INSTALAÇÕES','Entradas','Receitas','Aporte','Despesas','OPEX ADM', "OPEX OPERACIONAL",'CAPEX','Impostos']
def get_dados_graficos():
    import openpyxl
    arquivo_excel = 'Energy - Orçado x Realizado.xlsx'
    arquivo_excel = 'Energy - Orçado x Realizado.xlsx'
    wb = openpyxl.load_workbook(arquivo_excel, data_only=True)
    aba = wb.active
    ultima_linha = aba.max_row
    ultima_coluna = aba.max_column
    
    
    
    posicoes = []
    for i in range(ultima_coluna):
        celula = aba.cell(row=i+1, column=1).value
        posicoes.append(celula)
        posicoes.append(i+1)
    resumo = [] 
    dados = {}
    for j in descricoes:
        linha = j
        
        for i in range(len(posicoes)):
            if linha == posicoes[i]:
                pos = posicoes[i+1]
                # Extrai os dados
                realizadas = []
                orcado = [0]
                labels = []
                for k in range(ultima_linha):
                    cel = aba.cell(row=1, column=k+1)
                    if cel.value == "Realizado":
                        valor = aba.cell(row=pos, column=k+1).value
                        if valor is not None and valor < 0:
                            valor = valor * -1
                        realizadas.append(int(valor) if valor else 0)
                    if cel.value == "Orçado":
                        valor = aba.cell(row=pos, column=k+1).value
                        if valor == None:
                            valor = 0                        
                        if valor is not None and valor < 0:
                            valor = valor * -1
                        orcado.append(int(valor) if valor else 0)
                for k in range(len(realizadas)):
                    if k <= 2:
                        labels.append(f'{k+10}-24')
                    else:
                        labels.append(f'{k-2}-25')
                 
                dados[linha] = {"labels": labels, "realizadas": realizadas, "orcado": orcado}
       
    return dados

# Configuração da página
st.set_page_config(layout="wide", page_title="Projeto Pinhal")
# Executa ControlePinhal() apenas uma vez por sessão
if 'controlepinhal_rodado' not in st.session_state:
    get_dados_graficos()
    st.session_state['controlepinhal_rodado'] = True
# Lê o resumo
resumo_dict = {}
with open('resumo.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        resumo_dict[row['Categoria']] = row
        #print(row)
        

        
# Sidebar com logo e menu
with st.sidebar:
    st.image("logotipo.png", use_container_width=True)
    st.title("Painel Financeiro")
    st.markdown("Projeto de Iluminação Espírito Santo de Pinhal - SP")
    st.markdown("---")
    hoje = datetime.datetime.now()
    mes_atual_str = hoje.strftime('%m')
    st.markdown(f'<div style="font-size:1em; color:#123366; margin-bottom:0.5em;">Período: mês 10-24 até mês {mes_atual_str}-25 </div>', unsafe_allow_html=True)
    page = st.radio("Menu", ["Projeto Pinhal", "Resumo", "Gráficos", "Extratos Bancários"])
    
    
# Autenticação simples por senha
if 'autenticado' not in st.session_state:
    st.session_state['autenticado'] = False

if not st.session_state['autenticado']:
    st.markdown('<h2 style="color:#123366;">Acesso Restrito</h2>', unsafe_allow_html=True)
    senha = st.text_input('Digite a senha para acessar o painel:', type='password')
    if st.button('Entrar'):
        if senha == 'Pinhal2025':  # Troque por uma senha forte
            st.session_state['autenticado'] = True
            st.success('Acesso liberado!')
            #st.experimental_rerun()
        else:
            st.error('Senha incorreta!')
    st.stop()
st.markdown('<h1 style="font-size: 4em; color:#123366; margin-bottom:0.2em; text-align:center;">Projeto Pinhal</h1>', unsafe_allow_html=True)



if page == "Projeto Pinhal":
    # --- Bloco das linhas 133 a 227 (tabela HTML dos dados da planilha) ---
    arquivo_excel = 'Energy - Orçado x Realizado.xlsx'
    wb = openpyxl.load_workbook(arquivo_excel, data_only=True)
    aba = wb.active

    # Descobrir até qual coluna a primeira linha tem "Realizado"
    ultima_coluna = aba.max_column
    colunas_validas = []
    for col in range(1, ultima_coluna + 1):
        valor = aba.cell(row=1, column=col).value
        if valor == "Realizado":
            colunas_validas.append(col)
    if colunas_validas:
        col_fim = colunas_validas[-1]
    else:
        col_fim = ultima_coluna

    # Identificar colunas que NÃO começam com "Orçado"
    colunas_para_mostrar = []
    for j in range(1, col_fim + 1):
        valor = aba.cell(row=1, column=j).value
        if not (isinstance(valor, str) and valor.strip().lower().startswith("orçado")):
            colunas_para_mostrar.append(j)
    # Meses por extenso para linha 3
    meses_extenso = [
        "Out-24", "Nov-24", "Dez-24", "Jan-25", "Fev-25", "Mar-25",
        "Abr-25", "Mai-25", "Jun-25", "Jul-25", "Ago-25", "Set-25","Out-25", "Nov-25", "Dez-25"
    ]
    
        
    html = '''
        <style>
        .sticky-header th {
            position: sticky;
            top: 0;
            background: #e6f0fa;
            z-index: 2;
        }
        </style>
    '''
    html = '<h2 style="color:#123366; text-align:center;">Dados da Planilha Realizado 24/25</h2>'
    html += '<div style="overflow-x:auto;"><table style="border-collapse:collapse; width:100%;">'
    
    # Primeiro, encontra as linhas de APORTE e DESPESAS
    linha_aporte = None
    linha_despesas = None
    for i in range(1, 51):
        primeira_col = aba.cell(row=i, column=colunas_para_mostrar[0]).value
        nome = str(primeira_col).strip().upper() if isinstance(primeira_col, str) else ""
        if nome == "APORTE":
            linha_aporte = i
        elif nome == "DESPESAS":
            linha_despesas = i
            break
    
    for i in range(1, 61):  # Linhas 1 a 50
        if i in [1]:
            continue  # Pula a primeira linha

        primeira_coluna = aba.cell(row=i, column=colunas_para_mostrar[0]).value
        nome_linha = str(primeira_coluna).strip().upper() if isinstance(primeira_coluna, str) else ""

        linhas_destaque = ["ENTRADAS", "DESPESAS", "RECEITAS", "APORTE", "OPEX ADM", "OPEX OPERACIONAL", "CAPEX", "IMPOSTOS", "COFINS-2172"]
        is_destaque = nome_linha in linhas_destaque
        row_style = 'font-weight:bold;' if is_destaque else ''
        border_bottom = 'border-bottom:4px double #123366;' if is_destaque or i == 2 else ''

        html += f'<tr style="height:4px; {row_style}">'
        
        # Calcula a soma para a coluna 2 (índice 1) incluindo TODAS as colunas de valores
        if i != 2 and i != 3:  # Não é a linha de saldo bancário nem de meses
            soma_linha = 0
            
            # Se for a linha APORTE, soma todas as linhas ENTRE APORTE e DESPESAS (não incluindo APORTE nem DESPESAS)
            if nome_linha == "APORTE" and linha_aporte and linha_despesas:
                for linha_soma in range(linha_aporte + 1, linha_despesas):
                    for idx_soma, j_soma in enumerate(colunas_para_mostrar):
                        if idx_soma >= 2:  # A partir da coluna 3 (índice 2)
                            valor_soma = aba.cell(row=linha_soma, column=j_soma).value
                            if isinstance(valor_soma, (int, float)):
                                soma_linha += valor_soma if valor_soma else 0
            else:
                # Para outras linhas, soma apenas a própria linha
                for idx_soma, j_soma in enumerate(colunas_para_mostrar):
                    if idx_soma >= 2:  # A partir da coluna 3 (índice 2)
                        valor_soma = aba.cell(row=i, column=j_soma).value
                        if isinstance(valor_soma, (int, float)):
                            soma_linha += valor_soma if valor_soma else 0
        
        for idx, j in enumerate(colunas_para_mostrar):
            valor = aba.cell(row=i, column=j).value
            
            # Se for a linha APORTE, calcula a soma para CADA coluna individualmente
            if nome_linha == "APORTE" and linha_aporte and linha_despesas and idx >= 2 and i != 2 and i != 3:
                soma_coluna = 0
                for linha_soma in range(linha_aporte + 1, linha_despesas):
                    valor_soma = aba.cell(row=linha_soma, column=j).value
                    if isinstance(valor_soma, (int, float)):
                        soma_coluna += valor_soma if valor_soma else 0
                valor = soma_coluna
            # Substitui o valor da coluna 2 pela soma calculada (exceto linhas 2 e 3)
            elif idx == 1 and i != 2 and i != 3:
                valor = soma_linha

            # Linha 2: coluna 1 recebe "Saldo Bancário" em azul escuro e negrito
            if i == 2 and idx == 0:
                valor_formatado = '<span style="color:#123366; font-weight:bold;">Saldo Bancário</span>'
                style_color = ""
                align = "left"
                style = f'font-weight:bold; border-bottom:4px double #123366; padding:1px 3px; text-align:{align};'
            # Linha 2: demais colunas, valores em negrito, formata número com ponto
            elif i == 2:
                if isinstance(valor, (int, float)):
                    valor_formatado = f"<b>{int(valor):,}".replace(",", ".") + "</b>"
                    style_color = "color:red;" if valor < 0 else ""
                    align = "center"
                else:
                    valor_formatado = f"<b>{valor if valor is not None else ''}</b>"
                    style_color = ""
                    align = "center"
                style = f'font-weight:bold; border-bottom:4px double #123366; padding:1px 3px; text-align:{align};' + style_color
            # Linha 3: meses por extenso da lista fixa a partir da coluna 3
            elif i == 3:
                if idx >= 2 and (idx - 2) < len(meses_extenso):
                    valor_formatado = meses_extenso[idx - 2]
                else:
                    valor_formatado = ""
                style_color = ""
                align = "center"
                style = f'background:#e6f0fa; font-weight:bold; padding:0.5px 3px; text-align:{align};'
            # Demais linhas: formatação padrão
            elif isinstance(valor, (int, float)):
                valor_formatado = f"{int(valor):,}".replace(",", ".")
                style_color = "color:red;" if valor < 0 else ""
                align = "center"
                style = f'padding:1px 3px; text-align:{align};' + style_color
            else:
                valor_formatado = valor if valor is not None else ""
                style_color = ""
                if idx == 0 and nome_linha in ["ENTRADAS", "DESPESAS"]:
                    align = "center"
                else:
                    align = "left"
                style = f'padding:1px 3px; text-align:{align};' + style_color

            # Azul marinho e negrito para Entradas/Despesas na primeira coluna
            if idx == 0 and nome_linha in ["ENTRADAS", "DESPESAS"]:
                style += 'font-weight:bold; color:#123366;'

            # Negrito para linhas especiais (exceto Entradas/Despesas, que já estão acima)
            if is_destaque and not (nome_linha in ["ENTRADAS", "DESPESAS"] and idx == 0):
                style += 'font-weight:bold;'

            # Linha dupla em negrito para toda a linha de destaque ou linha 2
            if border_bottom and i != 2:
                style += border_bottom

            tag = 'th' if i == 3 else 'td'
            html += f'<{tag} style="border:1px solid #ccc; {style}">{valor_formatado}</{tag}>'
        html += '</tr>'
        
        
        
        
    html += '</table></div>'

    st.markdown(html, unsafe_allow_html=True)


if page == "Resumo":
    st.markdown('<h1 style="font-size:2em; color:#123366; margin-bottom:0.2em; text-align:center;">Resumo Financeiro</h1>', unsafe_allow_html=True)
    dados_graficos = get_dados_graficos()
    for nome in descricoes:
        st.markdown(f'<h1 style="text-align:left; color:#123366; font-size:1.3em; margin-bottom:0.2em;">{nome}</h1>', unsafe_allow_html=True)
        if nome in dados_graficos:
            realizadas = dados_graficos[nome]["realizadas"]
            orcado = dados_graficos[nome]["orcado"]
            soma_realizado = sum(realizadas)
            soma_orcado = sum(orcado)
            percentual = (soma_realizado / soma_orcado * 100) if soma_orcado != 0 else 0
            col1, col2, col3 = st.columns(3)
            
            if nome == "INSTALAÇÕES":
                with col1:
                    st.metric(f"{nome} Orçado no Período", f"{soma_orcado}")
                with col2:
                    st.metric(f"{nome} Realizado no Período", f"{soma_realizado}")
            else:
                with col1:
                    st.metric(f"{nome} Orçado no Período", f"R$ {soma_orcado:,.0f}".replace(",", "."))
                with col2:
                    st.metric(f"{nome} Realizado no Período", f"R$ {soma_realizado:,.0f}".replace(",", "."))
                
            with col3:
                st.metric(f"Execução {nome} (%)", f"{percentual:.1f}%")
        else:
            st.warning(f"Sem dados para {nome}")
        st.markdown('<hr style="height:2px;border:none;background:linear-gradient(90deg,rgba(18,51,102,0.18) 0%,rgba(46,196,182,0.18) 100%);border-radius:1px;margin:1px 0 1px 0;">', unsafe_allow_html=True)
        
        
elif page == "Gráficos":
    st.markdown('<h1 style="font-size:2em; color:#123366; margin-bottom:0.2em; text-align:center;">Gráficos Financeiro</h1>', unsafe_allow_html=True)
    dados_graficos = get_dados_graficos()
    
    for nome in descricoes:
        st.markdown(f'<h1 style="text-align:left; color:#123366; font-size:1.3em; margin-bottom:0.2em;">{nome}</h1>', unsafe_allow_html=True)
        labels = dados_graficos[nome]["labels"]
        realizadas = dados_graficos[nome]["realizadas"]
        orcado = dados_graficos[nome]["orcado"]

        # Calcula os resumos
        soma_realizado = sum(realizadas)
        soma_orcado = sum(orcado)
        diferenca = soma_realizado - soma_orcado
        percentual = (soma_realizado / soma_orcado * 100) if soma_orcado != 0 else 0

        # Mostra os resumos acima dos gráficos
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Soma Orçado", f"{soma_orcado:,.0f}".replace(",", "."))
        with col2:
            st.metric("Soma Realizado", f"{soma_realizado:,.0f}".replace(",", "."))
        with col3:
            st.metric("Diferença", f"{diferenca:,.0f}".replace(",", "."))
        with col4:
            st.metric("% Executada", f"{percentual:.1f}%")
        st.markdown('<hr style="height:2px;border:none;background:linear-gradient(90deg,rgba(18,51,102,0.18) 0%,rgba(46,196,182,0.18) 100%);border-radius:1px;margin:1px 0 1px 0;">', unsafe_allow_html=True)
        col_barra, col_linha = st.columns(2)
        with col_barra:
            st.markdown('<div style="text-align:center; font-weight:bold; color:#123366;">Barras</div>', unsafe_allow_html=True)
            fig_bar = go.Figure()
            fig_bar.add_trace(go.Bar(
                x=labels, y=realizadas, name='Realizadas', marker_color='navy',                
                text=[str(v) if nome == "INSTALAÇÕES" else (f"{v/1000:.1f}k" if abs(v) >= 1000 else str(v)) for v in realizadas],                                
                textposition='outside', textfont=dict(color='navy', size=30)
                ))
            fig_bar.add_trace(go.Bar(
                x=labels, y=orcado, name='Orçadas', marker_color='green', opacity=0.5
                ))
            
            
            fig_bar.update_layout(
                barmode='group',
                xaxis_title='Meses',
                yaxis_title='Valores',
                legend_title='Legenda',
                font=dict(size=60)  # aumenta o tamanho da fonte dos labels
            )
            
            
            st.plotly_chart(fig_bar, use_container_width=True)
        with col_linha:
            st.markdown('<div style="text-align:center; font-weight:bold; color:#2ec4b6;">Linha</div>', unsafe_allow_html=True)
            fig_line = go.Figure()
            fig_line.add_trace(go.Scatter(
                x=labels, y=realizadas, mode='lines+markers+text', name='Realizadas', line=dict(color='navy'),
                text=[f"{v/1000:.1f}k" if abs(v) >= 1000 else str(v) for v in realizadas], textposition='top center', textfont=dict(color='navy', size=20)
            ))
            fig_line.add_trace(go.Scatter(
                x=labels, y=orcado, mode='lines+markers', name='Orçadas', line=dict(color='green')
            ))
            
            
            fig_line.update_layout(
                xaxis_title='Meses',
                yaxis_title='Valores',
                legend_title='Legenda',
                font=dict(size=60)  # aumenta o tamanho da fonte dos labels
            )
            
            
            st.plotly_chart(fig_line, use_container_width=True)
    
        st.markdown('<hr style="height:2px;border:none;background:linear-gradient(90deg,rgba(18,51,102,0.18) 0%,rgba(46,196,182,0.18) 100%);border-radius:1px;margin:1px 0 1px 0;">', unsafe_allow_html=True)
        
        
elif page == "Extratos Bancários":
    st.markdown('<h1 style="font-size:2em; color:#123366; margin-bottom:0.2em; text-align:center;">Extratos Bancários - Upload, Visualização e Download</h1>', unsafe_allow_html=True)

    uploaded_pdf = st.file_uploader("Faça upload de um PDF de extrato bancário", type=["pdf"])
    if uploaded_pdf is not None:
        st.success(f"Arquivo '{uploaded_pdf.name}' enviado com sucesso!")
        st.download_button(f"Baixar {uploaded_pdf.name}", uploaded_pdf, file_name=uploaded_pdf.name)
    
    st.markdown('<h3 style="font-size:1em; color:#123366; margin-bottom:0.2em;">Arquivos disponíveis:</h3>', unsafe_allow_html=True)
    import glob
    pdfs = glob.glob("extratos/*.pdf")
    pdfs.sort(key=lambda x: os.path.basename(x).lower())
    for pdf in pdfs:
        nome_pdf = os.path.basename(pdf)
        with open(pdf, "rb") as f:
            st.download_button(f"Baixar {nome_pdf}", f, file_name=nome_pdf)



